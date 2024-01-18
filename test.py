import time
import keyboard
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
import PyPDF2
import re
import openpyxl

service = Service()
profile_path = r'C:\Users\drjuh\AppData\Local\Google\Chrome\User Data'  
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(f'--user-data-dir={profile_path}')
print("Cik neizlasītas vēstules jums ir?")
reps=int(input())

time.sleep(2)
driver = webdriver.Chrome(service=service, options=chrome_options)
#login ortus


login="Andrejs.Bistrovs"
password="Kolpo1234!"
driver.get('https://id2.rtu.lv/')
login_f=driver.find_element(By.NAME,"IDToken1")
login_f.click()
login_f.clear()
login_f.send_keys(login)
password_f=driver.find_element(By.ID,"IDToken2")
password_f.send_keys(password)
login_f=driver.find_element(By.NAME,"Login.Submit")
login_f.click()
#MAIL block
driver.get('https://mail.google.com/mail')
#parent=driver.getWindowHandle()




time.sleep(1)

search_input = driver.find_element(By.NAME, "q")

search_input.click()
search_input.send_keys("pievienojis atsauksmi")


search_input.send_keys(webdriver.Keys.RETURN)
for i in range (reps): 
    time.sleep(3)
    email_open= driver.find_element(By.CLASS_NAME, 'av')
    email_open.click()
    time.sleep(1)
    link_open= driver.find_element(By.LINK_TEXT,'uzdevuma iesniegumam')
    link_open.click()
    time.sleep(2)
    keyboard.press(['ctrl','p'])
    keyboard.release('ctrl')
    keyboard.release('p')
    time.sleep(2) 
    keyboard.press_and_release('enter')
    time.sleep(1)
    keyboard.write('crpdf.pdf')
    keyboard.press_and_release('enter')
    time.sleep(1)
    keyboard.press_and_release('tab')
    time.sleep(1)
    keyboard.press_and_release('enter')
    time.sleep(1)
    f= open(r'C:\Users\drjuh\Downloads\crpdf.pdf', "rb")
    pdf_file=PyPDF2.PdfReader(f)
    number_of_pages=len(pdf_file.pages)
    text=""

    for i in range (0,number_of_pages):
        text+=pdf_file.pages[i].extract_text()
    pos_after_nosaukums=text.find("Atvērts:")
    pos_before_nosaukums=text.find("E-studiju vide")
    subject_and_topic=text[pos_before_nosaukums+len(str(pos_before_nosaukums))+14:pos_after_nosaukums].rstrip()
    if len(subject_and_topic)>400:
        subject_and_topic=text[pos_before_nosaukums+len(str(pos_before_nosaukums))+14:text.find("Iesnieguma statuss")].rstrip()
    #print(subject_and_topic)
    match = re.search('[A-Z]', subject_and_topic)
    subject_and_topic=subject_and_topic[match.start():]
    #print(subject_and_topic)
    subject_and_topic=subject_and_topic.split(",")
    subject=subject_and_topic[0].strip()
    if("Datorgra" in subject):
        subject="Datorgrafikas un attēlu apstrādes pamati(1),23/24-R"
    topic=subject_and_topic[1].strip()
    print(subject)
    print(topic)

    if "Ieskaitı̄ts" in text:
            grade="Ieskaitı̄ts"
    else:
        pattern = r'\d+,\d+\s*\/\s*\d+'
        grade =str(re.findall(pattern, str(text)))
        grade=grade.strip("[]").replace("'", "")
        grade=grade.split('\\')[0].strip()
        grade=grade.replace('/','')
    print(grade)
    from openpyxl import Workbook, load_workbook 
    wb=load_workbook(r'C:\Users\drjuh\Downloads\uni_grades.xlsx')
    ws=wb.active
    ws=wb['Sheet1']
    max_row=ws.max_row
    Excelnumbers=["A","B",'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    done=0
    for i in range(1, max_row+1):
        if ws['A'+str(i)].value == subject:
            for j in Excelnumbers:
                if ws[j+str(i)].value is None:
                    ws[j+str(i)].value = topic
                    ws[j+str(i+1)].value = grade
                    done=1 
                    break
    if(done!=1): 
        ws['A'+str(max_row+1)].value=subject
        ws['B'+str(max_row+1)].value=topic
        ws['B'+str(max_row+2)].value=grade
            

    wb.save('uni_grades.xlsx')
    wb.close()
    keyboard.press(['ctrl','w'])
    keyboard.release('ctrl')
    keyboard.release('w')
    time.sleep(2)
    email_choose = driver.find_element(By.XPATH, '//body/div[7]/div[3]/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]')
    email_choose.click()
    time.sleep(2)
    f.close()
driver.quit()

